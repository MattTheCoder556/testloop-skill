#!/usr/bin/env python3
"""Shared helpers for the /testloop document chain.

Loads a Validation Test Matrix (the .xlsx or its .md twin) into plain dicts,
canonicalises outcome vocabulary, and applies the house filename convention so
every document in one iteration carries the same module token, timestamp and
version as the matrix it was built against.

Nothing here decides anything. It reads, names and styles; the three build
scripts own the arithmetic.
"""

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- style constants, matching the Validation Test Matrix workbook ---------
HEADER_FILL = "1F4E78"      # navy header band
HEADER_FONT = "FFFFFF"
GRID = "D9D9D9"             # thin cell borders
REGULATORY_FILL = "FCE4D6"  # peach row shading, Regulatory = Yes

# Outcome colours. Deliberately not the matrix's priority red: a failing test
# and a high-priority test are different things and must not look alike.
STATUS_STYLE = {
    "Pass":    {"color": "006100", "bold": False, "italic": False},
    "Fail":    {"color": "C00000", "bold": True,  "italic": False},
    "Blocked": {"color": "9C5700", "bold": True,  "italic": False},
    "Partial": {"color": "9C5700", "bold": False, "italic": False},
    "Not run": {"color": "808080", "bold": False, "italic": True},
}
STATUSES = list(STATUS_STYLE)

_STATUS_ALIASES = {
    "pass": "Pass", "passed": "Pass", "ok": "Pass", "p": "Pass",
    "fail": "Fail", "failed": "Fail", "f": "Fail",
    "blocked": "Blocked", "block": "Blocked", "b": "Blocked",
    "partial": "Partial", "partially passed": "Partial",
    "not run": "Not run", "notrun": "Not run", "not-run": "Not run",
    "untested": "Not run", "skipped": "Not run", "n/a": "Not run",
}

THIN = Side(style="thin", color=GRID)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


class LoopError(Exception):
    """A problem the operator has to resolve — printed without a traceback."""


def canon_status(value):
    """Map free-text outcome vocabulary onto the five canonical statuses."""
    if value is None:
        return "Not run"
    key = str(value).strip().lower()
    if not key:
        return "Not run"
    if key in _STATUS_ALIASES:
        return _STATUS_ALIASES[key]
    for status in STATUSES:
        if key == status.lower():
            return status
    raise LoopError(
        f"unknown status {value!r} — use one of: {', '.join(STATUSES)}"
    )


# --- matrix loading --------------------------------------------------------

_MATRIX_COLUMNS = {
    "test id": "id",
    "module / suite": "module",
    "module": "module",
    "user role": "role",
    "role": "role",
    "test case title": "title",
    "title": "title",
    "preconditions": "preconditions",
    "test steps": "steps",
    "steps": "steps",
    "expected result": "expected",
    "expected": "expected",
    "priority": "priority",
    "regulatory / compliance-critical": "regulatory",
    "regulatory": "regulatory",
}


def load_matrix(path, sheet=None):
    """Read a Validation Test Matrix into {module, sheet, rows:[...]}.

    Accepts the .xlsx (authoritative) or its .md twin. Row dicts carry
    id / module / role / title / preconditions / steps / expected / priority /
    regulatory, all as strings.
    """
    path = Path(path).expanduser()
    if not path.exists():
        raise LoopError(f"matrix not found: {path}")
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return _load_matrix_xlsx(path, sheet)
    if path.suffix.lower() in (".md", ".markdown"):
        return _load_matrix_md(path)
    raise LoopError(f"cannot read a matrix from {path.suffix} — give the .xlsx or .md")


def _load_matrix_xlsx(path, sheet):
    book = load_workbook(path, data_only=True)
    if sheet:
        if sheet not in book.sheetnames:
            raise LoopError(
                f"sheet {sheet!r} not in {path.name} — it has: "
                + ", ".join(book.sheetnames)
            )
        ws = book[sheet]
    elif len(book.sheetnames) == 1:
        ws = book[book.sheetnames[0]]
    else:
        raise LoopError(
            f"{path.name} holds {len(book.sheetnames)} sheets — name one with "
            f"--sheet: {', '.join(book.sheetnames)}"
        )

    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    index = {}
    for pos, name in enumerate(header):
        field = _MATRIX_COLUMNS.get(name)
        if field and field not in index:
            index[field] = pos
    missing = {"id", "title", "expected"} - set(index)
    if missing:
        raise LoopError(
            f"{path.name}!{ws.title} is missing column(s) {sorted(missing)} — "
            "is this a Validation Test Matrix sheet?"
        )

    rows = []
    for cells in ws.iter_rows(min_row=2, values_only=True):
        get = lambda f: str(cells[index[f]]).strip() if f in index and index[f] < len(cells) and cells[index[f]] is not None else ""
        if not get("id"):
            continue
        rows.append({f: get(f) for f in
                     ("id", "module", "role", "title", "preconditions",
                      "steps", "expected", "priority", "regulatory")})
    if not rows:
        raise LoopError(f"{path.name}!{ws.title} has no test-case rows")
    module = rows[0]["module"] or ws.title
    return {"module": module, "sheet": ws.title, "rows": rows, "path": str(path)}


_MD_HEADING = re.compile(r"^##\s+([A-Z][A-Z0-9]*(?:\s*/\s*[A-Z0-9]+)?-\d+)\s+[—-]\s+(.+?)\s*$")
_MD_ATTR = re.compile(r"^\|\s*\*\*(.+?)\*\*\s*\|\s*(.*?)\s*\|\s*$")


def _load_matrix_md(path):
    text = path.read_text(encoding="utf-8")
    module = ""
    head = re.search(r"^\*\*Module / Suite:\*\*\s*(.+?)\s*$", text, re.M)
    if head:
        module = head.group(1).strip()

    rows, current, section = [], None, None
    for line in text.splitlines():
        heading = _MD_HEADING.match(line)
        if heading:
            if current:
                rows.append(current)
            current = {"id": heading.group(1), "module": module, "role": "",
                       "title": heading.group(2), "preconditions": "",
                       "steps": [], "expected": "", "priority": "",
                       "regulatory": ""}
            section = None
            continue
        if current is None:
            continue
        attr = _MD_ATTR.match(line)
        if attr:
            field = _MATRIX_COLUMNS.get(attr.group(1).strip().lower())
            if field and field != "id":
                current[field] = attr.group(2).strip()
            continue
        if line.startswith("**Preconditions:**"):
            current["preconditions"] = line.split("**", 2)[2].lstrip(": ").strip()
            section = None
        elif line.startswith("**Test Steps:**"):
            section = "steps"
        elif line.startswith("**Expected Result:**"):
            current["expected"] = line.split("**", 2)[2].lstrip(": ").strip()
            section = None
        elif section == "steps":
            step = re.match(r"^\s*\d+[.)]\s+(.*\S)\s*$", line)
            if step:
                current["steps"].append(step.group(1))
    if current:
        rows.append(current)
    if not rows:
        raise LoopError(f"no test cases found in {path.name} — is this a matrix twin?")
    for row in rows:
        if isinstance(row["steps"], list):
            row["steps"] = " ".join(f"{n}) {s}" for n, s in enumerate(row["steps"], 1))
        row["module"] = row["module"] or module
    return {"module": module or rows[0]["module"], "sheet": module,
            "rows": rows, "path": str(path)}


# --- house naming ----------------------------------------------------------

_HOUSE_NAME = re.compile(
    r"^(?P<prefix>[A-Za-z]+)_(?P<token>[A-Za-z0-9]+)_"
    r"(?P<date>\d{4}-\d{2}-\d{2})_(?P<time>\d{4})_v(?P<version>\d+)$"
)


def parse_house_name(path):
    """Pull module token, timestamp and version out of a house filename.

    Returns None when the name does not follow the convention — the caller then
    has to be told the token and stamp rather than guessing them.
    """
    match = _HOUSE_NAME.match(Path(path).stem)
    if not match:
        return None
    return {
        "prefix": match.group("prefix"),
        "token": match.group("token"),
        "stamp": datetime.strptime(
            f"{match.group('date')} {match.group('time')}", "%Y-%m-%d %H%M"),
        "version": int(match.group("version")),
    }


def house_path(directory, prefix, token, stamp, version, ext):
    """<Prefix>_<Token>_<YYYY-MM-DD>_<HHMM>_v<N>.<ext> in `directory`."""
    name = (f"{prefix}_{token}_{stamp:%Y-%m-%d}_{stamp:%H%M}_v{version}.{ext}")
    return Path(directory).expanduser() / name


def derive_identity(args, matrix_path):
    """Settle token / stamp / version for this iteration.

    Defaults come from the matrix filename, so every derived document is named
    after the matrix it was built against and the whole iteration sorts
    together. Explicit flags win.
    """
    parsed = parse_house_name(matrix_path) or {}
    token = getattr(args, "token", None) or parsed.get("token")
    if not token:
        raise LoopError(
            f"cannot tell the module token from {Path(matrix_path).name} — pass "
            "--token (the <Module> field from the /validation registry, e.g. FormBuilder)"
        )
    stamp = parsed.get("stamp")
    if getattr(args, "stamp", None):
        stamp = datetime.strptime(args.stamp, "%Y-%m-%d %H%M")
    if stamp is None:
        raise LoopError(
            "cannot tell the iteration timestamp from the matrix filename — "
            "pass --stamp 'YYYY-MM-DD HHMM' and use the same value for every "
            "document in this iteration"
        )
    version = getattr(args, "version", None) or parsed.get("version") or 1
    return token, stamp, int(version)


# --- small output helpers --------------------------------------------------

def load_json(path):
    path = Path(path).expanduser()
    if not path.exists():
        raise LoopError(f"spec not found: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise LoopError(f"{path.name} is not valid JSON: {exc}") from exc


def md_cell(value):
    """Flatten a value for a Markdown table cell."""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        value = "; ".join(str(v) for v in value)
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def write_sheet(ws, headers, rows, widths, shade=None, cell_style=None):
    """Lay out a sheet in the workbook's house style.

    `shade` is a predicate on the row dict; `cell_style` a callable
    (row, field) -> Font or None for per-cell outcome colouring.
    """
    ws.append([h for h, _ in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    for row in rows:
        ws.append([md_cell(row.get(field, "")) if isinstance(row.get(field), (list, tuple))
                   else row.get(field, "") for _, field in headers])
        excel_row = ws.max_row
        shaded = shade(row) if shade else False
        for column, (_, field) in enumerate(headers, start=1):
            cell = ws.cell(row=excel_row, column=column)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            font = cell_style(row, field) if cell_style else None
            cell.font = font or Font(size=11)
            if shaded:
                cell.fill = PatternFill("solid", fgColor=REGULATORY_FILL)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
