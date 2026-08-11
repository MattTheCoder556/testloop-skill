#!/usr/bin/env python3
"""Audit the reconciliation back against the record, and confirm or refuse the run.

This is the fourth document, and the reason the loop is worth running. It does
**not** repeat the reconciliation's comparison — recomputing the same join with
the same inputs would agree with itself and prove nothing. It asks a different
question: *is the record strong enough to support the conclusion drawn from it?*

Three families of check:

  * **Evidence** — a Pass with nothing behind it is an assertion, not a result.
    A Fail with no reproduction is a rumour. Observed text that merely restates
    the expected result is a sign the row was filled in from the plan rather
    than from the screen.
  * **Integrity** — the rendered documents are recounted from the spec. If the
    results sheet or the reconciliation has been edited after generation, the
    numbers stop agreeing and this says so.
  * **Verdict** — the gate is recomputed independently and compared with the
    verdict the reconciliation printed.

Findings are Blocking or Advisory. Any Blocking finding refuses confirmation and
exits non-zero: the iteration does not close on a record that cannot carry it.

Usage:
    audit.py results.json --run-dir RUNDIR
    audit.py results.json --run-dir RUNDIR --evidence-root /path/to/screenshots
"""

import argparse
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

import lib
from lib import LoopError
from build_results import merge
from reconcile import assess

RESTATEMENT_RATIO = 0.90


def finding(code, severity, title, detail, cases=()):
    return {"code": code, "severity": severity, "title": title,
            "detail": detail, "cases": list(cases)}


# --- evidence checks -------------------------------------------------------

def check_evidence(rows, spec, evidence_root):
    out = []
    by_id = {e.get("id"): e for e in spec.get("results", []) if e.get("id")}

    unevidenced = [r["id"] for r in rows
                   if r["status"] == "Pass" and not r["evidence"]]
    if unevidenced:
        out.append(finding(
            "E1", "Blocking", "Passing cases with no evidence",
            "A Pass with no screenshot, log line or artefact behind it records that "
            "someone believes the case passed. Attach evidence or downgrade the status.",
            unevidenced))

    silent = [r["id"] for r in rows
              if r["status"] in ("Pass", "Fail", "Partial") and not r["observed"]]
    if silent:
        out.append(finding(
            "E2", "Blocking", "Outcome recorded with no observed result",
            "The status says what was concluded but not what was seen. Without the "
            "observation, nobody can re-derive the conclusion.", silent))

    missing_files = []
    for r in rows:
        for item in str(r["evidence"]).split("; "):
            item = item.strip()
            if not item or "://" in item:
                continue
            path = Path(item)
            if not path.is_absolute():
                path = evidence_root / item
            if not path.exists():
                missing_files.append(f"{r['id']} → {item}")
    if missing_files:
        out.append(finding(
            "E3", "Blocking", "Evidence referenced but not present",
            f"Resolved relative to `{evidence_root}`. A path that does not resolve is "
            "the same as no evidence once the run directory is filed or moved.",
            missing_files))

    no_repro = [r["id"] for r in rows if r["status"] == "Fail" and not r["_repro"]]
    if no_repro:
        out.append(finding(
            "E4", "Blocking", "Failures with no reproduction steps",
            "A failure nobody can reproduce cannot be fixed or retested, and will "
            "surface again in the next iteration as a fresh discovery.", no_repro))

    unexplained_blocks = [r["id"] for r in rows
                          if r["status"] in ("Blocked", "Partial")
                          and not (r["observed"] or r["notes"])]
    if unexplained_blocks:
        out.append(finding(
            "E5", "Blocking", "Blocked or partial cases with no reason given",
            "Say what blocked it. A block with no cause is indistinguishable from a "
            "case nobody got to.", unexplained_blocks))

    restated = []
    for r in rows:
        if r["status"] not in ("Pass", "Fail", "Partial") or not r["observed"]:
            continue
        ratio = SequenceMatcher(None, r["observed"].strip().lower(),
                                str(r["expected"]).strip().lower()).ratio()
        if ratio >= RESTATEMENT_RATIO:
            restated.append(f"{r['id']} ({ratio:.0%} identical)")
    if restated:
        out.append(finding(
            "E6", "Advisory", "Observed result restates the expected result",
            "Near-identical text usually means the row was completed from the plan "
            "rather than from the screen. Describe what appeared, in the words the "
            "product used.", restated))

    no_defect = [r["id"] for r in rows if r["status"] == "Fail" and not r["defect"]]
    if no_defect:
        out.append(finding(
            "E7", "Advisory", "Failures with no defect reference",
            "Raise the defect and record its id, or the failure lives only in this "
            "document and is lost at the next revision.", no_defect))

    thin_unplanned = [lib.md_cell(u.get("title", "untitled"))
                      for u in spec.get("unplanned", []) if not u.get("observed")]
    if thin_unplanned:
        out.append(finding(
            "E8", "Advisory", "Off-plan observations with no description",
            "An observation that is only a title cannot become a test case.",
            thin_unplanned))

    stale = [i for i in by_id if i not in {r["id"] for r in rows}]
    if stale:
        out.append(finding(
            "E9", "Blocking", "Recorded outcomes for cases the matrix does not contain",
            "Either the id is wrong, or the case was executed off-plan and belongs in "
            "`unplanned`.", stale))
    return out


# --- integrity checks ------------------------------------------------------

_TALLY = re.compile(r"^\|\s*(Pass|Fail|Blocked|Partial|Not run)\s*\|\s*(\d+)\s*\|")
_VERDICT = re.compile(r"^##\s+Verdict\s+[—-]\s+(PASS|HOLD)\s*$", re.M)


def check_integrity(rows, results_xlsx, reconciliation_md, criteria, blocking):
    out = []
    tally = {s: sum(1 for r in rows if r["status"] == s) for s in lib.STATUSES}

    if results_xlsx and results_xlsx.exists():
        try:
            ws = load_workbook(results_xlsx, data_only=True).active
            header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            id_col, status_col = header.index("test id"), header.index("status")
            rendered = {}
            for cells in ws.iter_rows(min_row=2, values_only=True):
                if cells[id_col]:
                    rendered[str(cells[id_col]).strip()] = str(cells[status_col] or "").strip()
            drift = [f"{r['id']}: sheet says {rendered.get(r['id'], '(absent)')!r}, "
                     f"record says {r['status']!r}"
                     for r in rows if rendered.get(r["id"]) != r["status"]]
            if drift:
                out.append(finding(
                    "I1", "Blocking", "Results sheet disagrees with the results record",
                    f"`{results_xlsx.name}` was recounted against the spec it was built "
                    "from. A mismatch means the sheet was edited by hand — regenerate it "
                    "from the corrected spec instead.", drift))
        except (ValueError, IndexError):
            out.append(finding(
                "I2", "Advisory", "Results sheet could not be recounted",
                f"`{results_xlsx.name}` has no recognisable Test ID / Status columns, so "
                "it was not cross-checked.", []))
    else:
        out.append(finding(
            "I3", "Advisory", "Results sheet not found",
            "The rendered results workbook was not available, so only the record was "
            "audited.", [str(results_xlsx) if results_xlsx else "—"]))

    if reconciliation_md and reconciliation_md.exists():
        text = reconciliation_md.read_text(encoding="utf-8")
        stated = {m.group(1): int(m.group(2)) for m in
                  (_TALLY.match(line) for line in text.splitlines()) if m}
        drift = [f"{s}: reconciliation says {stated[s]}, recount says {tally[s]}"
                 for s in lib.STATUSES if s in stated and stated[s] != tally[s]]
        if not stated:
            out.append(finding(
                "I4", "Advisory", "Reconciliation tally could not be read",
                f"No outcome tally table found in `{reconciliation_md.name}`.", []))
        elif drift:
            out.append(finding(
                "I5", "Blocking", "Reconciliation counts disagree with the record",
                f"`{reconciliation_md.name}` was recounted from the results spec.", drift))

        match = _VERDICT.search(text)
        recomputed = "PASS" if not blocking else "HOLD"
        if not match:
            out.append(finding(
                "I6", "Advisory", "No verdict found in the reconciliation",
                f"`{reconciliation_md.name}` states no verdict to check.", []))
        elif match.group(1) != recomputed:
            out.append(finding(
                "I7", "Blocking", "Stated verdict does not match the criteria",
                f"The reconciliation says {match.group(1)}; recomputing the gate from "
                f"the record gives {recomputed}.",
                [c["name"] for c in blocking] or ["all criteria met"]))
    else:
        out.append(finding(
            "I8", "Blocking", "Reconciliation not found",
            "There is nothing to audit. Run reconcile.py first.",
            [str(reconciliation_md) if reconciliation_md else "—"]))
    return out


# --- document ---------------------------------------------------------------

def write_md(path, matrix, spec, rows, findings, criteria, blocking,
             sources, confirmed):
    tally = {s: sum(1 for r in rows if r["status"] == s) for s in lib.STATUSES}
    blockers = [f for f in findings if f["severity"] == "Blocking"]
    advisories = [f for f in findings if f["severity"] == "Advisory"]

    out = []
    out.append(f"# Audit & Confirmation — {matrix['module']}\n")
    out.append(f"**Module / Suite:** {matrix['module']}  ")
    for label, key in (("Matrix (plan)", "matrix"), ("Results (record)", "results"),
                       ("Reconciliation", "reconciliation")):
        if sources.get(key):
            out.append(f"**{label}:** `{Path(sources[key]).name}`  ")
    if spec.get("environment"):
        out.append(f"**Environment:** {spec['environment']}  ")
    out.append("")
    out.append("This pass does not re-compare the plan with the record — the "
               "reconciliation already did that, and repeating it would only agree with "
               "itself. It asks whether the record is strong enough to carry the verdict: "
               "whether every outcome has evidence behind it, and whether the rendered "
               "documents still match the record they were generated from.\n")

    out.append(f"## Confirmation — {'CONFIRMED' if confirmed else 'REFUSED'}\n")
    if confirmed:
        out.append(f"The record supports the reconciliation's verdict. "
                   f"{len(advisories)} advisory finding(s) remain and do not block.\n")
    else:
        out.append(f"**Not confirmed.** {len(blockers)} blocking finding(s). The verdict "
                   "in the reconciliation is not supported by the record as it stands — "
                   "fix the record and regenerate the chain, or re-run the affected "
                   "cases.\n")

    out.append("| | |")
    out.append("|---|---|")
    out.append(f"| **Cases** | {len(rows)} |")
    out.append("| **Outcome** | "
               + ", ".join(f"{tally[s]} {s.lower()}" for s in lib.STATUSES if tally[s])
               + " |")
    out.append(f"| **Gate** | {'PASS' if not blocking else 'HOLD'}"
               f" ({len(blocking)} blocking criterion(s) unmet) |")
    out.append(f"| **Audit** | {len(blockers)} blocking, {len(advisories)} advisory |")
    out.append("")

    out.append("## Findings\n")
    if not findings:
        out.append("None. Every outcome carries an observation and evidence, the rendered "
                   "documents recount correctly, and the verdict follows from the "
                   "criteria.\n")
    else:
        out.append("| Code | Severity | Finding | Cases |")
        out.append("|---|---|---|---|")
        for f in findings:
            out.append(f"| {f['code']} | {'**' + f['severity'] + '**' if f['severity'] == 'Blocking' else f['severity']} "
                       f"| {f['title']} | {len(f['cases']) or '—'} |")
        out.append("")
        for f in findings:
            out.append(f"### {f['code']} — {f['title']} ({f['severity']})\n")
            out.append(f["detail"] + "\n")
            for case in f["cases"]:
                out.append(f"- {case}")
            if f["cases"]:
                out.append("")

    out.append("## Gate criteria, recomputed\n")
    out.append("Recomputed here from the record, independently of the reconciliation:\n")
    out.append("| Criterion | Met | Blocking |")
    out.append("|---|---|---|")
    for c in criteria:
        out.append(f"| {c['name']} | {'Yes' if c['ok'] else '**No**'} | "
                   f"{'Yes' if c['blocking'] else 'No'} |")
    out.append("")

    out.append("## What closing this iteration would mean\n")
    if confirmed and not blocking:
        out.append("- The matrix's cases were all executed and all passed, on the "
                   "environment named above.")
        out.append("- Each outcome is backed by an observation and evidence that exists.")
        out.append("- The four documents in this iteration agree with each other.\n")
        out.append("It does **not** mean the module is correct — only that it behaves as "
                   "the matrix says it should. Anything the matrix does not cover is "
                   "untested, and the plan gaps listed in the reconciliation are the "
                   "known part of that.\n")
    else:
        out.append("Nothing closes yet. Carry the blocking findings and the unmet criteria "
                   "into the next iteration: revise the matrix with `/validation` where the "
                   "plan was wrong, re-run the affected cases where the record was thin, "
                   "and rebuild this chain at the next version.\n")

    out.append("---\n")
    out.append("| Sign-off | Name | Date |")
    out.append("|---|---|---|")
    out.append("| Executed by | | |")
    out.append("| Reviewed by | | |")
    out.append("| Approved by | | |")
    out.append("")
    out.append("_Generated. Regenerate it from the record rather than editing it — an "
               "edited audit audits nothing._\n")

    Path(path).write_text("\n".join(out) + "\n", encoding="utf-8")


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("spec", help="the results spec JSON the chain was built from")
    ap.add_argument("--matrix", help="matrix .xlsx or .md (overrides spec.matrix)")
    ap.add_argument("--sheet")
    ap.add_argument("--run-dir", metavar="DIR",
                    help="iteration directory holding the Results and Reconciliation files")
    ap.add_argument("--results-xlsx", help="explicit path to the rendered results workbook")
    ap.add_argument("--reconciliation", help="explicit path to the reconciliation .md")
    ap.add_argument("--out", help="explicit .md path for the audit")
    ap.add_argument("--evidence-root", help="directory evidence paths resolve against "
                                            "(default: the spec's directory)")
    ap.add_argument("--token")
    ap.add_argument("--stamp", help="iteration timestamp 'YYYY-MM-DD HHMM'")
    ap.add_argument("--version", type=int)
    ap.add_argument("--require-full-coverage", action="store_true")
    args = ap.parse_args(argv)

    spec = lib.load_json(args.spec)
    spec_dir = Path(args.spec).expanduser().parent
    matrix_path = args.matrix or spec.get("matrix")
    if not matrix_path:
        raise LoopError("no matrix given — set 'matrix' in the spec or pass --matrix")
    matrix_path = (Path(matrix_path) if Path(matrix_path).is_absolute()
                   else spec_dir / matrix_path)
    matrix = lib.load_matrix(matrix_path, args.sheet or spec.get("sheet"))
    rows, _ = merge(matrix, spec, True)
    criteria, blocking = assess(rows, spec.get("unplanned", []),
                                args.require_full_coverage)

    token, stamp, version = lib.derive_identity(args, matrix_path)
    run_dir = Path(args.run_dir).expanduser() if args.run_dir else spec_dir
    results_xlsx = (Path(args.results_xlsx).expanduser() if args.results_xlsx
                    else lib.house_path(run_dir, "Results", token, stamp, version, "xlsx"))
    reconciliation = (Path(args.reconciliation).expanduser() if args.reconciliation
                      else lib.house_path(run_dir, "Reconciliation", token, stamp,
                                          version, "md"))
    audit_md = (Path(args.out).expanduser() if args.out
                else lib.house_path(run_dir, "Audit", token, stamp, version, "md"))
    evidence_root = (Path(args.evidence_root).expanduser() if args.evidence_root
                     else run_dir)

    findings = (check_evidence(rows, spec, evidence_root)
                + check_integrity(rows, results_xlsx, reconciliation, criteria, blocking))
    blockers = [f for f in findings if f["severity"] == "Blocking"]
    confirmed = not blockers and not blocking

    audit_md.parent.mkdir(parents=True, exist_ok=True)
    write_md(audit_md, matrix, spec, rows, findings, criteria, blocking,
             {"matrix": matrix["path"], "results": results_xlsx,
              "reconciliation": reconciliation}, confirmed)

    print(f"md: {audit_md}")
    print(f"findings: {len(blockers)} blocking, "
          f"{len(findings) - len(blockers)} advisory")
    for f in blockers:
        print(f"  [{f['code']}] {f['title']}: "
              + (", ".join(str(c) for c in f["cases"][:6]) or "—"))
    print(f"confirmation: {'CONFIRMED' if confirmed else 'REFUSED'}")
    return 0 if confirmed else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except LoopError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(2)
